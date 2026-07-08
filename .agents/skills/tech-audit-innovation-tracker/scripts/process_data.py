#!/usr/bin/env python3
"""
IT Innovation Tracker — Data Processor

Ingests raw project management data and optional financial data,
classifies each item as maintenance or innovation, computes metrics,
and outputs the canonical tracker JSON.

Usage:
    python3 process_data.py --input data.csv [options]
    python3 process_data.py --input data.xlsx --sheet "Projects" [options]
    python3 process_data.py --input data.json [options]

Options:
    --input FILE          Project/task data (CSV, XLSX, JSON)
    --financials FILE     Optional P&L or budget data (CSV, XLSX)
    --sheet NAME          Sheet name for Excel files (default: first sheet)
    --period TEXT          Period label (default: auto-detect or "Current")
    --currency CODE       ISO currency code (default: USD)
    --metric-type TYPE    cost | effort | hybrid (default: auto-detect)
    --metric-unit UNIT    $ | hours | SP | FTE | tasks (default: auto-detect)
    --industry TEXT        Industry for benchmarks (default: other)
    --org-name TEXT        Organization name
    --output FILE          Output JSON path (default: tracker_output.json)
    --history FILE         Optional historical periods JSON for trend data
"""

import argparse
import csv
import json
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

# ─── Classification Keywords ────────────────────────────────────────────────

MAINTENANCE_HIGH = [
    "bug", "bugfix", "hotfix", "defect", "incident",
    "patch", "security patch", "cve", "vulnerability",
    "upgrade", "version upgrade", "dependency update",
    "compliance", "audit", "regulatory", "sox", "gdpr",
    "backup", "disaster recovery", "dr test",
    "monitoring", "alerting", "health check",
    "license renewal", "subscription renewal",
    "support ticket", "helpdesk", "service request",
    "infrastructure maintenance", "server maintenance",
    "performance tuning", "tech debt", "technical debt",
    "eol", "end of life", "decommission",
    "sla", "uptime", "availability", "not working", "not showing"
]

MAINTENANCE_MEDIUM = [
    "migration", "optimization", "testing", "documentation",
    "training", "integration fix", "sync issue", "refactor",
    "remove", "review", "landing page", "wrong", "clean up", "figure out",
    "fix", "not being", "protect", "determine", "re-implement",
    "qa", "ensure", "hardening", "move", "error message"
]

INNOVATION_HIGH = [
    "new feature", "new functionality", "new capability",
    "new product", "new service", "new platform",
    "greenfield", "net-new", "from scratch",
    "proof of concept", "poc", "prototype", "mvp",
    "r&d", "research", "experiment",
    "ai", "ml", "machine learning",
    "digital transformation", "modernization",
    "new integration", "new api", "new connector",
    "ux redesign", "ui overhaul",
    "new market", "expansion", "growth initiative",
    "analytics platform", "data lake",
    "automation",
]

INNOVATION_MEDIUM = [
    "platform migration", "cloud migration",
    "redesign", "process improvement", "dashboard", "feature",
    "implement", "functionality", "build", "create", "integration",
    "needs new", "allow"
]

# Source labels that map directly
LABEL_MAP = {
    "bug": "maintenance",
    "defect": "maintenance",
    "task": None,  # ambiguous
    "story": None,
    "feature": "innovation",
    "enhancement": "innovation",
    "epic": None,
    "spike": "innovation",
    "maintenance": "maintenance",
    "support": "maintenance",
    "incident": "maintenance",
    "change request": None,
    "improvement": None,
}

# Industry benchmarks
INDUSTRY_BENCHMARKS = {
    "financial_services": (70, 30),
    "healthcare": (65, 35),
    "retail": (55, 45),
    "technology": (47, 53),
    "manufacturing": (70, 30),
    "government": (80, 20),
    "telecom": (65, 35),
    "energy": (75, 25),
    "education": (70, 30),
    "media": (50, 50),
    "other": (65, 35),
}


def classify_item(name, source_label=None):
    """Classify a work item as maintenance or innovation.

    Returns (category, confidence, reason).
    """
    name_lower = name.lower().strip()

    # 1. Check explicit source label
    if source_label:
        label_lower = source_label.lower().strip()
        mapped = LABEL_MAP.get(label_lower)
        if mapped:
            return mapped, "high", f"Source label '{source_label}' maps to {mapped}"

    # 2. High-confidence keyword match
    for kw in MAINTENANCE_HIGH:
        if kw in name_lower:
            return "maintenance", "high", f"Matched maintenance keyword: '{kw}'"

    for kw in INNOVATION_HIGH:
        if kw in name_lower:
            return "innovation", "high", f"Matched innovation keyword: '{kw}'"

    # 3. Medium-confidence keyword match
    for kw in MAINTENANCE_MEDIUM:
        if kw in name_lower:
            return "maintenance", "medium", f"Matched medium-confidence keyword: '{kw}' (defaulting maintenance)"

    for kw in INNOVATION_MEDIUM:
        if kw in name_lower:
            return "innovation", "medium", f"Matched medium-confidence keyword: '{kw}'"

    # 4. Default to maintenance (conservative)
    return "maintenance", "low", "No keyword match - defaulting to maintenance (conservative)"


def parse_value(raw):
    """Parse a numeric value from various formats: $150,000 / 150000 / 150k."""
    if raw is None:
        return 0.0
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip().replace(",", "").replace("$", "").replace("€", "").replace("£", "")
    s = s.lower()
    multiplier = 1
    if s.endswith("k"):
        multiplier = 1000
        s = s[:-1]
    elif s.endswith("m"):
        multiplier = 1_000_000
        s = s[:-1]
    try:
        return float(s) * multiplier
    except ValueError:
        return 0.0


def detect_columns(headers):
    """Auto-detect column roles from header names."""
    mapping = {"name": None, "value": None, "type": None, "team": None}
    headers_lower = [h.lower().strip() for h in headers]

    name_candidates = ["project", "name", "title", "epic", "task", "summary", "description", "item"]
    value_candidates = ["cost", "budget", "spend", "value", "amount", "estimate",
                        "hours", "story points", "sp", "points", "effort", "fte"]
    type_candidates = ["type", "category", "label", "issue type", "issuetype",
                       "classification", "kind", "tag", "work type"]
    team_candidates = ["team", "department", "group", "squad", "org", "unit",
                       "assignee team", "cost center"]

    for i, h in enumerate(headers_lower):
        for c in name_candidates:
            if c in h and mapping["name"] is None:
                mapping["name"] = i
        for c in value_candidates:
            if c in h and mapping["value"] is None:
                mapping["value"] = i
        for c in type_candidates:
            if c in h and mapping["type"] is None:
                mapping["type"] = i
        for c in team_candidates:
            if c in h and mapping["team"] is None:
                mapping["team"] = i

    return mapping


def read_csv(filepath):
    """Read a CSV file and return headers + rows."""
    rows = []
    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        headers = next(reader)
        for row in reader:
            if any(cell.strip() for cell in row):
                rows.append(row)
    return headers, rows


def read_excel(filepath, sheet_name=None):
    """Read an Excel file using openpyxl. Returns headers + rows."""
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl required for Excel files. Install: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    data = list(ws.iter_rows(values_only=True))
    if not data:
        return [], []
    headers = [str(h) if h else f"col_{i}" for i, h in enumerate(data[0])]
    rows = [[str(c) if c is not None else "" for c in row] for row in data[1:] if any(c is not None for c in row)]
    return headers, rows


def read_json_input(filepath):
    """Read a JSON array of objects. Returns headers + rows."""
    with open(filepath, "r") as f:
        data = json.load(f)
    if isinstance(data, list) and len(data) > 0:
        headers = list(data[0].keys())
        rows = [[str(item.get(h, "")) for h in headers] for item in data]
        return headers, rows
    return [], []


def generate_recommendations(summary, top_maint, top_innov, history=None):
    """Generate actionable recommendations based on the data."""
    recs = []
    pct = summary["innovation_pct"]

    if pct > 40:
        recs.append({
            "type": "positive",
            "icon": "✦",
            "text": f"Strong innovation ratio at {pct:.1f}%. "
                    f"Verify maintenance coverage is sufficient — your top maintenance "
                    f"cost driver is \"{top_maint[0]['name']}\" at {top_maint[0]['value']:,.0f}."
                    if top_maint else f"Strong innovation ratio at {pct:.1f}%."
        })
        if len(top_maint) > 1:
            smallest = top_maint[-1]
            recs.append({
                "type": "positive",
                "icon": "→",
                "text": f"Consider automating \"{smallest['name']}\" ({smallest['value']:,.0f}) "
                        f"to free up even more capacity for innovation."
            })
    elif pct > 25:
        recs.append({
            "type": "warning",
            "icon": "⚠",
            "text": f"Innovation at {pct:.1f}% is below the 40% target. "
                    f"Review maintenance items for automation or outsourcing opportunities."
        })
        if top_maint:
            recs.append({
                "type": "warning",
                "icon": "→",
                "text": f"Top maintenance drain: \"{top_maint[0]['name']}\" "
                        f"({top_maint[0]['value']:,.0f}). Evaluate if this can be reduced."
            })
    else:
        recs.append({
            "type": "critical",
            "icon": "🚨",
            "text": f"Critical: Only {pct:.1f}% on innovation. The organization is in a "
                    f"keep-the-lights-on trap. Executive intervention needed."
        })
        recs.append({
            "type": "critical",
            "icon": "→",
            "text": "Immediate action: Audit all maintenance items. Retire zombie applications, "
                    "automate operations, consolidate redundant systems."
        })

    if history and len(history) > 1:
        first = history[0]["innovation_pct"]
        last = history[-1]["innovation_pct"]
        trend = "improving" if last > first else ("declining" if last < first else "stable")
        recs.append({
            "type": "trend",
            "icon": "📈",
            "text": f"Trend is {trend}: innovation moved from {first:.0f}% to {last:.0f}% "
                    f"over {len(history)} periods."
        })

    return recs


def process(args):
    """Main processing pipeline."""
    # ─── Read input ──────────────────────────────────────────────────────
    ext = Path(args.input).suffix.lower()
    if ext == ".csv":
        headers, rows = read_csv(args.input)
    elif ext in (".xlsx", ".xls", ".xlsm"):
        headers, rows = read_excel(args.input, args.sheet)
    elif ext == ".json":
        headers, rows = read_json_input(args.input)
    else:
        print(f"ERROR: Unsupported file type: {ext}", file=sys.stderr)
        sys.exit(1)

    if not rows:
        print("ERROR: No data rows found in input file.", file=sys.stderr)
        sys.exit(1)

    # ─── Detect columns ─────────────────────────────────────────────────
    col_map = detect_columns(headers)
    if col_map["name"] is None:
        print(f"WARNING: Could not auto-detect name column from headers: {headers}", file=sys.stderr)
        col_map["name"] = 0  # fallback to first column

    # ─── Detect metric type ──────────────────────────────────────────────
    metric_type = args.metric_type
    metric_unit = args.metric_unit
    if not metric_type and col_map["value"] is not None:
        val_header = headers[col_map["value"]].lower()
        if any(w in val_header for w in ["cost", "budget", "spend", "amount", "$"]):
            metric_type = "cost"
            metric_unit = metric_unit or "$"
        elif any(w in val_header for w in ["hour"]):
            metric_type = "effort"
            metric_unit = metric_unit or "hours"
        elif any(w in val_header for w in ["point", "sp"]):
            metric_type = "effort"
            metric_unit = metric_unit or "SP"
        else:
            metric_type = "effort"
            metric_unit = metric_unit or "tasks"
    metric_type = metric_type or "effort"
    metric_unit = metric_unit or "tasks"

    # ─── Classify items ──────────────────────────────────────────────────
    items = []
    for i, row in enumerate(rows):
        name = row[col_map["name"]].strip() if col_map["name"] < len(row) else f"Item {i+1}"
        value = parse_value(row[col_map["value"]]) if col_map["value"] is not None and col_map["value"] < len(row) else 1.0
        source_label = row[col_map["type"]].strip() if col_map["type"] is not None and col_map["type"] < len(row) else None
        team = row[col_map["team"]].strip() if col_map["team"] is not None and col_map["team"] < len(row) else None

        category, confidence, reason = classify_item(name, source_label)

        items.append({
            "id": f"item-{i+1:03d}",
            "name": name,
            "category": category,
            "value": value,
            "unit": metric_unit,
            "team": team or "",
            "source_label": source_label or "",
            "confidence": confidence,
            "classification_reason": reason,
            "flagged": confidence == "low",
            "notes": "",
        })

    # ─── Compute summary ─────────────────────────────────────────────────
    total = sum(it["value"] for it in items)
    maint_val = sum(it["value"] for it in items if it["category"] == "maintenance")
    innov_val = sum(it["value"] for it in items if it["category"] == "innovation")
    maint_count = sum(1 for it in items if it["category"] == "maintenance")
    innov_count = sum(1 for it in items if it["category"] == "innovation")
    innov_pct = (innov_val / total * 100) if total > 0 else 0
    maint_pct = 100 - innov_pct

    health = "healthy" if innov_pct > 40 else ("at_risk" if innov_pct > 25 else "critical")
    health_labels = {
        "healthy": "Strong innovation investment",
        "at_risk": "Below target — room for improvement",
        "critical": "Maintenance-heavy — innovation starved",
    }

    summary = {
        "total_value": round(total, 2),
        "maintenance_value": round(maint_val, 2),
        "innovation_value": round(innov_val, 2),
        "maintenance_pct": round(maint_pct, 1),
        "innovation_pct": round(innov_pct, 1),
        "item_count": len(items),
        "maintenance_count": maint_count,
        "innovation_count": innov_count,
        "health_status": health,
        "health_label": health_labels[health],
        "flagged_items_count": sum(1 for it in items if it["flagged"]),
    }

    # ─── By team ─────────────────────────────────────────────────────────
    teams = {}
    for it in items:
        t = it["team"] or "Unassigned"
        if t not in teams:
            teams[t] = {"team": t, "total_value": 0, "maintenance_value": 0, "innovation_value": 0}
        teams[t]["total_value"] += it["value"]
        teams[t][f"{it['category']}_value"] += it["value"]

    by_team = []
    for t in sorted(teams.values(), key=lambda x: x["total_value"], reverse=True):
        tv = t["total_value"]
        t["maintenance_pct"] = round(t["maintenance_value"] / tv * 100, 1) if tv else 0
        t["innovation_pct"] = round(t["innovation_value"] / tv * 100, 1) if tv else 0
        t["total_value"] = round(t["total_value"], 2)
        t["maintenance_value"] = round(t["maintenance_value"], 2)
        t["innovation_value"] = round(t["innovation_value"], 2)
        by_team.append(t)

    # ─── Top lists ───────────────────────────────────────────────────────
    maint_items = sorted([it for it in items if it["category"] == "maintenance"], key=lambda x: x["value"], reverse=True)[:5]
    innov_items = sorted([it for it in items if it["category"] == "innovation"], key=lambda x: x["value"], reverse=True)[:5]

    top_maintenance = [{"name": it["name"], "value": round(it["value"], 2), "pct_of_total": round(it["value"] / total * 100, 1)} for it in maint_items]
    top_innovation = [{"name": it["name"], "value": round(it["value"], 2), "pct_of_total": round(it["value"] / total * 100, 1)} for it in innov_items]

    # ─── History ─────────────────────────────────────────────────────────
    history = []
    if args.history and os.path.exists(args.history):
        with open(args.history) as f:
            history = json.load(f)

    # Append current period to history
    history.append({
        "period": args.period,
        "innovation_pct": round(innov_pct, 1),
        "maintenance_pct": round(maint_pct, 1),
        "total_value": round(total, 2),
    })

    # ─── Industry benchmark ──────────────────────────────────────────────
    industry = args.industry or "other"
    bench = INDUSTRY_BENCHMARKS.get(industry, INDUSTRY_BENCHMARKS["other"])
    diff = round(innov_pct - bench[1], 1)
    if diff > 5:
        comp = "above_benchmark"
        comp_label = f"Your innovation ratio is {diff:.0f} points above your industry average ({bench[1]}%)"
    elif diff < -5:
        comp = "below_benchmark"
        comp_label = f"Your innovation ratio is {abs(diff):.0f} points below your industry average ({bench[1]}%)"
    else:
        comp = "at_benchmark"
        comp_label = f"Your innovation ratio is in line with your industry average ({bench[1]}%)"

    industry_benchmark = {
        "industry": industry,
        "typical_maintenance_pct": bench[0],
        "typical_innovation_pct": bench[1],
        "comparison": comp,
        "comparison_label": comp_label,
    }

    # ─── Recommendations ─────────────────────────────────────────────────
    recs = generate_recommendations(summary, top_maintenance, top_innovation, history if len(history) > 1 else None)

    # ─── Classification stats ────────────────────────────────────────────
    classification_stats = {
        "auto_classified": len(items),
        "user_overridden": 0,
        "flagged_for_review": summary["flagged_items_count"],
        "high_confidence": sum(1 for it in items if it["confidence"] == "high"),
        "medium_confidence": sum(1 for it in items if it["confidence"] == "medium"),
        "low_confidence": sum(1 for it in items if it["confidence"] == "low"),
    }

    # ─── Assemble output ─────────────────────────────────────────────────
    output = {
        "schema_version": "1.0",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "meta": {
            "title": f"IT Portfolio Analysis — {args.period}",
            "period": args.period,
            "currency": args.currency if metric_type == "cost" else None,
            "metric_type": metric_type,
            "metric_unit": metric_unit,
            "org_name": args.org_name or None,
            "industry": industry,
            "notes": None,
        },
        "summary": summary,
        "items": items,
        "by_team": by_team if any(t["team"] != "Unassigned" for t in by_team) else [],
        "top_maintenance": top_maintenance,
        "top_innovation": top_innovation,
        "history": history if len(history) > 1 else [],
        "recommendations": recs,
        "industry_benchmark": industry_benchmark,
        "classification_stats": classification_stats,
    }

    # ─── Write output ────────────────────────────────────────────────────
    output_path = args.output or "tracker_output.json"
    with open(output_path, "w") as f:
        json.dump(output, f, indent=2)

    print(f"✓ Processed {len(items)} items → {output_path}")
    print(f"  Innovation: {innov_pct:.1f}% ({innov_count} items, {innov_val:,.0f} {metric_unit})")
    print(f"  Maintenance: {maint_pct:.1f}% ({maint_count} items, {maint_val:,.0f} {metric_unit})")
    print(f"  Health: {health.upper()} — {health_labels[health]}")
    if summary["flagged_items_count"]:
        print(f"  ⚠ {summary['flagged_items_count']} items flagged for review (low confidence)")

    return output


def main():
    parser = argparse.ArgumentParser(description="IT Innovation Tracker — Data Processor")
    parser.add_argument("--input", required=True, help="Project data file (CSV, XLSX, JSON)")
    parser.add_argument("--financials", help="Optional P&L / budget file")
    parser.add_argument("--sheet", help="Sheet name for Excel files")
    parser.add_argument("--period", default="Current", help="Period label")
    parser.add_argument("--currency", default="USD", help="Currency code")
    parser.add_argument("--metric-type", choices=["cost", "effort", "hybrid"], help="Metric type")
    parser.add_argument("--metric-unit", help="Metric unit ($, hours, SP, FTE, tasks)")
    parser.add_argument("--industry", default="other", help="Industry for benchmarks")
    parser.add_argument("--org-name", help="Organization name")
    parser.add_argument("--output", default="tracker_output.json", help="Output JSON path")
    parser.add_argument("--history", help="Historical periods JSON file")
    args = parser.parse_args()
    process(args)


if __name__ == "__main__":
    main()
